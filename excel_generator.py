import io
from datetime import date, timedelta
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
LOCATION_FILL = PatternFill("solid", start_color="D6E4F0")
LOCATION_FONT = Font(bold=True, name="Arial", size=10)
CELL_FONT = Font(name="Arial", size=9)
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _week_ranges(year: int, month: int) -> list[tuple[date, date]]:
    """Return (monday, sunday) pairs covering all days in the month."""
    first = date(year, month, 1)
    last = date(year, month, monthrange(year, month)[1])
    # start from the Monday of the week containing the 1st
    start = first - timedelta(days=first.weekday())
    weeks = []
    while start <= last:
        end = start + timedelta(days=6)
        weeks.append((start, end))
        start += timedelta(days=7)
    return weeks


def _tab_name(week_start: date, week_end: date) -> str:
    return f"{week_start.strftime('%d%b')}-{week_end.strftime('%d%b')}"


def _resolve_location(loc: str, ordered: list[str]) -> str:
    """Map a log's location to its best match in the ordered list.
    Handles prefix mismatches like 'CCW' stored in DB vs 'CCW1' in setorder.
    Mirrors the resolution used by the daily report so the Excel matches rows.
    """
    if loc in ordered:
        return loc
    for candidate in ordered:
        if candidate.startswith(loc) or loc.startswith(candidate):
            return candidate
    return loc


def _has_description(log: dict) -> bool:
    """True if the log carries an actual activity description.

    Entries without one say nothing about what happened on site, so they are
    left out of the export entirely rather than rendered as a blank row. None,
    "" and whitespace-only all count as absent.
    """
    return bool(str(log.get("description") or "").strip())


def _style_cell(cell, font=None, fill=None, alignment=None, border=True):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = BORDER


def generate_monthly_excel(
    group_id: str,
    year: int,
    month: int,
    logs: list[dict],
    locations: list[str],
) -> bytes:
    """
    Build a monthly Excel workbook.
    - One sheet per week
    - Rows: only days that have logged activity, grouped by location
    - Columns: Day | Date | Main Location | Sub Location | Description | Manpower

    Every row carries a real description. Entries without one are dropped, and
    days with no activity are not written at all rather than padded with "-".
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    weeks = _week_ranges(year, month)

    # Index logs by (resolved location, date) for quick lookup. Resolve the
    # stored main_location to its canonical name (e.g. "CCW" -> "CCW1") so rows
    # match the ordered location list — otherwise every cell renders "-".
    log_index: dict[tuple[str, str], list[dict]] = {}
    for log in logs:
        if not _has_description(log):
            continue
        resolved = _resolve_location(log["main_location"], locations)
        key = (resolved, log["log_date"])
        log_index.setdefault(key, []).append(log)

    # Include any locations present in the logs but missing from the ordered
    # list, so unmapped entries still appear (mirrors the daily report).
    locations = list(locations)
    for resolved, _log_date in log_index:
        if resolved not in locations:
            locations.append(resolved)

    for week_start, week_end in weeks:
        ws = wb.create_sheet(title=_tab_name(week_start, week_end))
        ws.sheet_view.showGridLines = False

        # Column widths
        col_widths = [8, 12, 22, 30, 55, 20]
        col_headers = ["Day", "Date", "Main Location", "Sub Location", "Description / Activity", "Manpower"]
        for i, (w, h) in enumerate(zip(col_widths, col_headers), start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
            cell = ws.cell(row=1, column=i, value=h)
            _style_cell(cell, font=HEADER_FONT, fill=HEADER_FILL,
                        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

        ws.row_dimensions[1].height = 30

        current_row = 2
        week_days = [week_start + timedelta(days=d) for d in range(7)]

        for location in locations:
            wrote_any = False
            for day in week_days:
                if day.month != month:
                    continue  # spill-over days from the neighbouring month

                day_label = f"{DAY_NAMES[day.weekday()]} ({day.strftime('%d/%m')})"

                # Only days with actual logged activity get a row. Previously a
                # row was written for every location on every day, padded with
                # "-" — on July 2026 that was 3337 filler rows against 313 real
                # ones, burying the report in blanks.
                for entry in log_index.get((location, day.isoformat()), []):
                    cells_data = [
                        day_label,
                        day.strftime("%d %b %Y"),
                        entry.get("main_location", location),
                        entry.get("sub_location", ""),
                        entry.get("description", ""),
                        entry.get("manpower", ""),
                    ]
                    for col, val in enumerate(cells_data, start=1):
                        cell = ws.cell(row=current_row, column=col, value=val)
                        _style_cell(cell, font=CELL_FONT,
                                    alignment=Alignment(vertical="top", wrap_text=True))
                    ws.row_dimensions[current_row].height = 35
                    current_row += 1
                    wrote_any = True

            # Separator only after a location that actually produced rows,
            # otherwise absent locations leave a run of blank rows behind.
            if wrote_any:
                current_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Column groups: (group_label, [(header, field, width), ...])
_DWALL_GROUPS = [
    ("Panel Identification", [
        ("Report Date",  "report_date",        14),
        ("Engineer",     "engineer_initials",   10),
        ("Entry No.",    "entry_number",        10),
        ("Panel No.",    "panel_number",        14),
        ("Group",        "panel_group",          8),
    ]),
    ("Panel Specifications", [
        ("Panel Size",       "panel_size",        16),
        ("Guide Wall Level", "guide_wall_level",  18),
        ("Cut Off Level",    "cut_off_level",     16),
        ("Design Toe Level", "design_toe_level",  18),
        ("Design Depth",     "design_depth",      14),
        ("Final Depth",      "final_depth",       12),
        ("Rock Hit",         "rock_hit",          10),
    ]),
    ("Excavation",   [("Start", "excavation_start", 22), ("End", "excavation_end", 22)]),
    ("Koden Check",  [("Start", "koden_start",      22), ("End", "koden_end",      22)]),
    ("Desanding",    [("Start", "desanding_start",  22), ("End", "desanding_end",  22)]),
    ("Water Stop",   [("Start", "water_stop_start", 22), ("End", "water_stop_end", 22)]),
    ("Rebar Cage",   [("Start", "rebar_cage_start", 22), ("End", "rebar_cage_end", 22)]),
    ("Tremie Pipe",  [("Start", "tremie_pipe_start",22), ("End", "tremie_pipe_end",22)]),
    ("Casting",      [("Start", "casting_start",    22), ("End", "casting_end",    22)]),
    ("Volumes", [
        ("Theo Vol",    "theo_volume",   14),
        ("Actual Vol",  "actual_volume", 14),
        ("Overbreak %", "overbreak_pct", 12),
    ]),
    ("Other", [
        ("Downtime", "downtime", 35),
        ("Notes",    "notes",    25),
    ]),
]

GROUP_FILL   = PatternFill("solid", start_color="2E75B6")
GROUP_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=9)
SUBHDR_FILL  = PatternFill("solid", start_color="1F4E79")
SUBHDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=9)


def _format_downtime(downtime) -> str:
    if not downtime:
        return ""
    if isinstance(downtime, list):
        lines = []
        for entry in downtime:
            if isinstance(entry, dict):
                d = entry.get("date", "")
                s = entry.get("start", "")
                e = entry.get("end", "")
                r = entry.get("reason", "")
                lines.append(f"{d} {s}–{e}: {r}".strip(": "))
            else:
                lines.append(str(entry))
        return "\n".join(lines)
    return str(downtime)


def generate_dwall_excel(panels: list[dict]) -> bytes:
    """Generate a D-Wall / Barrette panel tracker Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Panel Tracker"
    ws.sheet_view.showGridLines = False

    # Flatten groups into ordered column list
    columns: list[tuple[str, str, int]] = []   # (sub_header, field, width)
    group_spans: list[tuple[str, int, int]] = [] # (label, start_col, end_col) 1-based
    col = 1
    for group_label, cols in _DWALL_GROUPS:
        group_spans.append((group_label, col, col + len(cols) - 1))
        for sub_header, field, width in cols:
            columns.append((sub_header, field, width))
            col += 1

    # Row 1 — group headers (merged)
    for label, start_col, end_col in group_spans:
        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1,   end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=label)
        _style_cell(cell, font=GROUP_FONT, fill=GROUP_FILL,
                    alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

    # Row 2 — sub-headers + column widths
    for i, (sub_header, _field, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
        cell = ws.cell(row=2, column=i, value=sub_header)
        _style_cell(cell, font=SUBHDR_FONT, fill=SUBHDR_FILL,
                    alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 25

    # Data rows start at row 3
    for row_i, panel in enumerate(panels, start=3):
        row_height = 20
        for col_i, (_sub_header, field, _w) in enumerate(columns, start=1):
            raw = panel.get(field, "")
            if field == "downtime":
                val = _format_downtime(raw)
                if val and "\n" in val:
                    row_height = max(row_height, 15 * (val.count("\n") + 1))
            else:
                val = raw or ""
            cell = ws.cell(row=row_i, column=col_i, value=val)
            _style_cell(cell, font=CELL_FONT,
                        alignment=Alignment(vertical="top", wrap_text=True))
        ws.row_dimensions[row_i].height = row_height

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
