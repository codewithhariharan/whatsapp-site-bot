"""Excel export generation — verify it produces valid, openable workbooks."""
import io
from datetime import date

import pytest
from openpyxl import load_workbook

import excel_generator as xls


class TestWeekRanges:
    def test_weeks_cover_the_whole_month(self):
        weeks = xls._week_ranges(2026, 6)
        first_start, _ = weeks[0]
        _, last_end = weeks[-1]
        # The covering weeks must bracket June 1st and June 30th.
        assert first_start <= date(2026, 6, 1)
        assert last_end >= date(2026, 6, 30)

    def test_each_range_is_a_full_monday_to_sunday_week(self):
        for start, end in xls._week_ranges(2026, 6):
            assert start.weekday() == 0   # Monday
            assert (end - start).days == 6


class TestFormatDowntime:
    def test_empty_downtime_is_blank(self):
        assert xls._format_downtime([]) == ""
        assert xls._format_downtime(None) == ""

    def test_structured_downtime_is_rendered(self):
        out = xls._format_downtime(
            [{"date": "21/02/26", "start": "14:00hrs", "end": "15:30hrs",
              "reason": "Equipment fault"}]
        )
        assert "21/02/26" in out
        assert "Equipment fault" in out


def _log(description, location="Zone1", sub="GL-A", manpower="2", day=15):
    return {
        "main_location": location, "sub_location": sub,
        "description": description, "manpower": manpower,
        "log_date": date(2026, 6, day).isoformat(),
    }


def _all_cell_values(data: bytes) -> list:
    wb = load_workbook(io.BytesIO(data))
    return [c.value for ws in wb.worksheets for row in ws.iter_rows() for c in row]


def _data_rows(data: bytes) -> list[list]:
    """Every written row below the header, excluding blank separator rows."""
    wb = load_workbook(io.BytesIO(data))
    rows = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            values = [c.value for c in row]
            if any(v is not None for v in values):
                rows.append(values)
    return rows


class TestMonthlyExcel:
    def test_produces_openable_workbook(self):
        data = xls.generate_monthly_excel("g1", 2026, 6, [_log("Rebar")], ["Zone1"])
        assert isinstance(data, bytes) and len(data) > 0
        wb = load_workbook(io.BytesIO(data))
        assert len(wb.sheetnames) >= 1  # one sheet per week

    def test_log_with_a_description_is_included(self):
        data = xls.generate_monthly_excel("g1", 2026, 6, [_log("Rebar works")], ["Zone1"])
        assert "Rebar works" in _all_cell_values(data)


class TestDescriptionRequired:
    """Entries with no description say nothing about site activity, so they are
    left out of the export rather than written as a blank row."""

    @pytest.mark.parametrize("description", [None, "", "   ", "\n", "\t "])
    def test_log_without_a_description_is_excluded(self, description):
        data = xls.generate_monthly_excel(
            "g1", 2026, 6, [_log(description, sub="UNIQUE-SUB")], ["Zone1"])
        # Its other fields must not appear either — the whole row is dropped.
        assert "UNIQUE-SUB" not in _all_cell_values(data)

    def test_missing_description_key_is_excluded(self):
        log = _log("x", sub="UNIQUE-SUB")
        del log["description"]
        data = xls.generate_monthly_excel("g1", 2026, 6, [log], ["Zone1"])
        assert "UNIQUE-SUB" not in _all_cell_values(data)

    def test_described_logs_survive_alongside_undescribed_ones(self):
        logs = [_log("Real activity", sub="KEEP"), _log(None, sub="DROP")]
        values = _all_cell_values(
            xls.generate_monthly_excel("g1", 2026, 6, logs, ["Zone1"]))
        assert "KEEP" in values
        assert "DROP" not in values

    def test_dropping_the_only_entry_leaves_no_row_behind(self):
        data = xls.generate_monthly_excel("g1", 2026, 6, [_log(None)], ["Zone1"])
        assert _data_rows(data) == []

    def test_location_only_referenced_by_undescribed_logs_is_not_invented(self):
        # Unordered locations are appended from the logs; a dropped log must not
        # drag its location into the sheet.
        logs = [_log(None, location="GhostZone")]
        assert "GhostZone" not in _all_cell_values(
            xls.generate_monthly_excel("g1", 2026, 6, logs, ["Zone1"]))


class TestNoPlaceholderRows:
    """Days without activity are omitted, not padded with "-".

    The sheet used to write a row for every location on every day of every
    week; on July 2026 that produced 3337 filler rows around 313 real ones.
    """

    def test_days_without_logs_produce_no_rows(self):
        rows = _data_rows(
            xls.generate_monthly_excel("g1", 2026, 6, [_log("Rebar", day=15)], ["Zone1"]))
        assert len(rows) == 1
        assert rows[0][4] == "Rebar"

    def test_no_dash_filler_anywhere(self):
        data = xls.generate_monthly_excel("g1", 2026, 6, [_log("Rebar")], ["Zone1"])
        assert "-" not in _all_cell_values(data)

    def test_ordered_location_with_no_activity_is_absent(self):
        # Locations come from /setorder, but an empty one must not fill the
        # sheet with a week of dashes.
        data = xls.generate_monthly_excel(
            "g1", 2026, 6, [_log("Rebar", location="Zone1")], ["Zone1", "Zone2"])
        values = _all_cell_values(data)
        assert "Zone1" in values
        assert "Zone2" not in values

    def test_month_with_no_logs_at_all_is_empty_but_valid(self):
        data = xls.generate_monthly_excel("g1", 2026, 6, [], ["Zone1"])
        wb = load_workbook(io.BytesIO(data))
        assert len(wb.sheetnames) >= 1          # week tabs still present
        assert wb.worksheets[0].cell(1, 1).value == "Day"   # headers intact
        assert _data_rows(data) == []

    def test_spillover_days_from_other_months_are_not_written(self):
        # A week tab spans a month boundary; only in-month days may appear.
        data = xls.generate_monthly_excel("g1", 2026, 6, [_log("Rebar", day=1)], ["Zone1"])
        for row in _data_rows(data):
            assert row[1].startswith("01 Jun")

    def test_multiple_locations_keep_one_separator_each(self):
        logs = [_log("A", location="Zone1"), _log("B", location="Zone2")]
        data = xls.generate_monthly_excel("g1", 2026, 6, logs, ["Zone1", "Zone2"])
        rows = _data_rows(data)
        assert [r[4] for r in rows] == ["A", "B"]
        # One blank separator between them, none trailing into a run.
        wb = load_workbook(io.BytesIO(data))
        ws = [w for w in wb.worksheets if w.max_row > 1][0]
        blanks = sum(1 for r in ws.iter_rows(min_row=2)
                     if all(c.value is None for c in r))
        assert blanks <= 2


class TestFullExcel:
    """/excel2 — every log on one sheet, oldest first."""

    def _logs(self):
        return [
            {"main_location": "Zone2", "sub_location": "b", "description": "June work",
             "manpower": "2", "log_date": "2026-06-15", "logged_at": "2026-06-15T09:00"},
            {"main_location": "Zone1", "sub_location": "a", "description": "April work",
             "manpower": "1", "log_date": "2026-04-20", "logged_at": "2026-04-20T09:00"},
            {"main_location": "Zone1", "sub_location": "c", "description": "July work",
             "manpower": "3", "log_date": "2026-07-27", "logged_at": "2026-07-27T09:00"},
        ]

    def test_single_sheet(self):
        wb = load_workbook(io.BytesIO(
            xls.generate_full_excel(self._logs(), ["Zone1", "Zone2"])))
        assert wb.sheetnames == ["All Logs"]

    def test_rows_are_chronological_oldest_first(self):
        rows = _data_rows(xls.generate_full_excel(self._logs(), ["Zone1", "Zone2"]))
        assert [r[4] for r in rows] == ["April work", "June work", "July work"]

    def test_keeps_the_monthly_column_layout(self):
        wb = load_workbook(io.BytesIO(
            xls.generate_full_excel(self._logs(), ["Zone1"])))
        header = [c.value for c in wb["All Logs"][1]]
        assert header == ["Day", "Date", "Main Location", "Sub Location",
                          "Description / Activity", "Manpower"]

    def test_day_and_date_columns_are_populated(self):
        rows = _data_rows(xls.generate_full_excel(self._logs(), ["Zone1", "Zone2"]))
        assert rows[0][0] == "Mon (20/04)"   # 20 Apr 2026 is a Monday
        assert rows[0][1] == "20 Apr 2026"

    def test_within_a_day_follows_setorder(self):
        logs = [
            {"main_location": "Zone2", "description": "second", "log_date": "2026-06-15",
             "sub_location": "", "manpower": "", "logged_at": "2026-06-15T08:00"},
            {"main_location": "Zone1", "description": "first", "log_date": "2026-06-15",
             "sub_location": "", "manpower": "", "logged_at": "2026-06-15T09:00"},
        ]
        rows = _data_rows(xls.generate_full_excel(logs, ["Zone1", "Zone2"]))
        assert [r[4] for r in rows] == ["first", "second"]

    def test_unordered_locations_sort_last_but_are_kept(self):
        logs = [
            {"main_location": "Ghost", "description": "unordered", "log_date": "2026-06-15",
             "sub_location": "", "manpower": "", "logged_at": "2026-06-15T08:00"},
            {"main_location": "Zone1", "description": "ordered", "log_date": "2026-06-15",
             "sub_location": "", "manpower": "", "logged_at": "2026-06-15T09:00"},
        ]
        rows = _data_rows(xls.generate_full_excel(logs, ["Zone1"]))
        assert [r[4] for r in rows] == ["ordered", "unordered"]

    def test_undescribed_logs_are_excluded(self):
        logs = self._logs() + [
            {"main_location": "Zone1", "sub_location": "NOPE", "description": None,
             "manpower": "", "log_date": "2026-05-01", "logged_at": "2026-05-01T09:00"},
        ]
        assert "NOPE" not in _all_cell_values(
            xls.generate_full_excel(logs, ["Zone1"]))

    def test_no_dash_filler(self):
        assert "-" not in _all_cell_values(
            xls.generate_full_excel(self._logs(), ["Zone1", "Zone2"]))

    def test_header_is_frozen_for_scrolling(self):
        wb = load_workbook(io.BytesIO(
            xls.generate_full_excel(self._logs(), ["Zone1"])))
        assert wb["All Logs"].freeze_panes == "A2"

    def test_empty_logs_still_produce_a_valid_workbook(self):
        data = xls.generate_full_excel([], ["Zone1"])
        wb = load_workbook(io.BytesIO(data))
        assert wb.sheetnames == ["All Logs"]
        assert _data_rows(data) == []


class TestDwallExcel:
    def test_produces_workbook_with_panel_tracker_sheet(self):
        panels = [{"panel_number": "CN284A", "entry_number": "Ent-2",
                   "downtime": []}]
        data = xls.generate_dwall_excel(panels)
        wb = load_workbook(io.BytesIO(data))
        assert wb.active.title == "Panel Tracker"
